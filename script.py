import binascii
import decimal
import os
from collections import defaultdict
from decimal import Decimal

import base58
from bitcoinrpc.authproxy import AuthServiceProxy, JSONRPCException
from Crypto.Hash import SHA256, RIPEMD160
from openpyxl.workbook import Workbook

from dotenv import load_dotenv

load_dotenv()


def connect_to_node():
    rpc_user = os.getenv('RPC_USER')
    rpc_password = os.getenv('RPC_PASSWORD')
    rpc_host = os.getenv('RPC_HOST', '127.0.0.1')
    rpc_port = os.getenv('RPC_PORT', '8332')

    rpc_connection = AuthServiceProxy(f"http://{rpc_user}:{rpc_password}@{rpc_host}:{rpc_port}")
    return rpc_connection

def get_block_transactions(rpc_connection, block_hash):
    try:
        block = rpc_connection.getblock(block_hash, 2)
        return block['tx']
    except JSONRPCException as e:
        print(f"Error fetching block: {e}")
        return []

def pubkey_to_address(pubkey: str) -> str:
    sha256_result = SHA256.new(bytes.fromhex(pubkey)).digest()
    ripemd160 = RIPEMD160.new()
    ripemd160.update(sha256_result)
    ripemd160_result = ripemd160.digest()
    versioned_payload = b"\x00" + ripemd160_result
    checksum = SHA256.new(SHA256.new(versioned_payload).digest()).digest()[:4]
    binary_address = versioned_payload + checksum
    bitcoin_address = base58.b58encode(binary_address).decode("utf-8")
    return bitcoin_address

def script_to_p2sh_address(script: str, mainnet=True) -> str:
    script_bytes = binascii.unhexlify(script)
    sha256 = SHA256.new(script_bytes).digest()
    ripemd160 = RIPEMD160.new(sha256).digest()
    version_byte = b"\x05" if mainnet else b"\xc4"
    payload = version_byte + ripemd160
    checksum = SHA256.new(SHA256.new(payload).digest()).digest()[:4]
    return base58.b58encode(payload + checksum).decode()

def script_to_p2pkh_address(script: str, mainnet=True) -> str:
    pubkey_hash = script.split()[2]
    version_byte = b"\x00" if mainnet else b"\x6f"
    payload = version_byte + binascii.unhexlify(pubkey_hash)
    checksum = SHA256.new(SHA256.new(payload).digest()).digest()[:4]
    return base58.b58encode(payload + checksum).decode()

def derive_address(script_pub_key: dict, script_pub_key_asm: str) -> str:
    script_type = script_pub_key.get("type", "")

    if "address" in script_pub_key:
        return script_pub_key["address"]

    if "addresses" in script_pub_key and script_pub_key["addresses"]:
        return script_pub_key["addresses"][0]

    if script_type == "pubkey":
        pubkey = script_pub_key_asm.split()[0]
        return pubkey_to_address(pubkey)

    if script_type == "pubkeyhash":
        return script_to_p2pkh_address(script_pub_key["hex"])

    if script_type == "scripthash":
        return script_to_p2sh_address(script_pub_key["hex"])

    if script_type == "multisig":
        return script_to_p2sh_address(script_pub_key["hex"])

    if script_type in ["witness_v0_keyhash", "witness_v0_scripthash"]:
        return script_pub_key.get("address", "")  # Bech32 address should be provided

    # fallback
    if "OP_CHECKSIG" in script_pub_key_asm:
        pubkey = script_pub_key_asm.split()[0]
        return pubkey_to_address(pubkey)
    elif "OP_CHECKMULTISIG" in script_pub_key_asm:
        return script_to_p2sh_address(script_pub_key["hex"])

    # return a placeholder if address isn't interpretable
    return f"UNKNOWN_{script_type}"


def batch_getrawtransaction(rpc_connection, txids):
    commands = [["getrawtransaction", txid, True] for txid in txids]
    try:
        return rpc_connection.batch_(commands)
    except JSONRPCException as e:
        print(f"Error in batch getrawtransaction: {e}")
        return []


def btc_to_satoshis(btc_value):
    return int(Decimal(str(btc_value)) * Decimal('100000000'))


def analyze_block(rpc_connection, block_hash, output_file, target_address=None):
    transactions = get_block_transactions(rpc_connection, block_hash)

    # Collect all input txids
    input_txids = []
    for tx in transactions:
        for vin in tx['vin']:
            if 'txid' in vin:
                input_txids.append(vin['txid'])

    # Batch fetch all input transactions
    input_txs = batch_getrawtransaction(rpc_connection, input_txids)
    input_tx_dict = {tx['txid']: tx for tx in input_txs if tx is not None}

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Transaction Details"
    ws1.append(['Input Transaction', 'Input Index', 'Output Transaction', 'Output Index', 'Incoming Address',
                'Outgoing Address', 'Amount (satoshis)'])

    ws2 = wb.create_sheet(title="Transaction Hashes")
    ws2.append(['Transaction Hash'])

    address_stats = defaultdict(lambda: {'balance_in': Decimal('0'), 'balance_out': Decimal('0'), 'tx_count': 0})

    # Use a set to store unique transaction inputs and outputs
    unique_txs = set()

    # Create a list to store transactions involving the target address
    target_address_txs = []

    for tx in transactions:
        ws2.append([tx['txid']])

        for vin_index, vin in enumerate(tx['vin']):
            if 'coinbase' in vin:
                # Handle coinbase transaction
                for vout_index, vout in enumerate(tx['vout']):
                    amount_satoshis = btc_to_satoshis(vout['value'])
                    output_address = derive_address(vout['scriptPubKey'], vout['scriptPubKey'].get('asm', ''))

                    unique_key = ('Coinbase', vin_index, tx['txid'], vout_index)
                    if unique_key not in unique_txs:
                        unique_txs.add(unique_key)
                        ws1.append([
                            'Coinbase',
                            vin_index,
                            tx['txid'],
                            vout_index,
                            'Coinbase (Newly generated coins)',
                            output_address,
                            amount_satoshis
                        ])

                    address_stats[output_address]['balance_in'] += Decimal(amount_satoshis)
                    address_stats[output_address]['tx_count'] += 1

                    if output_address == target_address:
                        target_address_txs.append({
                            'type': 'output',
                            'txid': tx['txid'],
                            'vout_index': vout_index,
                            'amount': amount_satoshis,
                            'address': output_address
                        })


            elif 'txid' in vin:

                input_tx = input_tx_dict.get(vin['txid'])
                if input_tx:
                    input_vout = input_tx['vout'][vin['vout']]
                    input_address = derive_address(input_vout['scriptPubKey'],
                                                   input_vout['scriptPubKey'].get('asm', ''))
                    input_amount = btc_to_satoshis(input_vout['value'])

                    # Only update stats for unique inputs
                    unique_input_key = (vin['txid'], vin['vout'])

                    if unique_input_key not in unique_txs:
                        unique_txs.add(unique_input_key)
                        address_stats[input_address]['balance_out'] += Decimal(input_amount)
                        address_stats[input_address]['tx_count'] += 1
                        if input_address == target_address:
                            target_address_txs.append({
                                'type': 'input',
                                'txid': tx['txid'],
                                'vin_index': vin_index,
                                'amount': input_amount,
                                'address': input_address
                            })

                    for vout_index, vout in enumerate(tx['vout']):
                        amount_satoshis = btc_to_satoshis(vout['value'])
                        output_address = derive_address(vout['scriptPubKey'], vout['scriptPubKey'].get('asm', ''))
                        unique_key = (tx['txid'], vout_index)

                        if unique_key not in unique_txs:
                            unique_txs.add(unique_key)
                            ws1.append([
                                vin['txid'],
                                vin['vout'],
                                tx['txid'],
                                vout_index,
                                input_address,
                                output_address,
                                amount_satoshis
                            ])

                            # Only update stats for unique outputs

                            address_stats[output_address]['balance_in'] += Decimal(amount_satoshis)
                            address_stats[output_address]['tx_count'] += 1

                            if output_address == target_address:
                                target_address_txs.append({
                                    'type': 'output',
                                    'txid': tx['txid'],
                                    'vout_index': vout_index,
                                    'amount': amount_satoshis,
                                    'address': output_address
                                })
                else:
                    print(f"Warning: Input transaction {vin['txid']} not found")

    ws3 = wb.create_sheet(title="Address Statistics")
    ws3.append(['Address', 'Balance In (satoshis)', 'Balance Out (satoshis)', 'Number of Transactions'])

    for address, stats in address_stats.items():
        ws3.append([
            address,
            int(stats['balance_in']),
            int(stats['balance_out']),
            stats['tx_count']
        ])

    if target_address:
        ws4 = wb.create_sheet(title=f"Transactions for {target_address}")
        ws4.append(['Type', 'Transaction Hash', 'Index', 'Amount (satoshis)', 'Address'])
        for tx in target_address_txs:
            ws4.append([
                tx['type'],
                tx['txid'],
                tx.get('vin_index', tx.get('vout_index')),
                tx['amount'],
                tx['address']
            ])

    wb.save(output_file)
    print(f"Excel file '{output_file}' has been created with the transaction details and statistics.")


def main():
    # Replace these with your actual RPC credentials and node information
    rpc_connection = connect_to_node()

    output_file = "block_transactions.xlsx"
    target_address = os.getenv('TARGET_ADDRESS')  # Add this to your .env file if you want to analyze a specific address

    analyze_block(rpc_connection, os.getenv('BLOCK_HASH', ''), output_file, target_address)


if __name__ == "__main__":
    main()
